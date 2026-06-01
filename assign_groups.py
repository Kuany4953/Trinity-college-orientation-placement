import pandas as pd
import random
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import sys


# ─────────────────────────────────────────
# CONFIGURATION — edit these each year
# ─────────────────────────────────────────
NUM_GROUPS        = 49
MIN_GROUP_SIZE    = 12
MAX_GROUP_SIZE    = 16

MAX_NE_PER_GROUP          = 10   # Northeast total
MAX_MACT_PER_GROUP        = 6    # MA + CT specifically
MAX_NON_NE_DOM_PER_GROUP  = 4    # Non-NE domestic
# International rule: each group must have either 0 internationals, OR
# between MIN_INTL_PER_GROUP and MAX_INTL_PER_GROUP (inclusive).
# A group with exactly 1 international student is NOT allowed.
MIN_INTL_PER_GROUP        = 2
MAX_INTL_PER_GROUP        = 3

NORTHEAST_STATES = {"CT", "MA", "ME", "NH", "NY", "RI", "VT"}
MACT_STATES      = {"MA", "CT"}

RANDOM_SEED = 42  # change to get a different valid arrangement

# ─────────────────────────────────────────
# COLUMN NAME MAP — update to match your actual Excel headers
# ─────────────────────────────────────────
COL_STATE       = "Address 1 Geomarket"
COL_COUNTRY     = "Citizenship (Primary)"
COL_HIGH_SCHOOL = "School 1 Name"
COL_SPORT       = "Sport 1 Name"
COL_POSSE       = "Posse"
COL_SEX         = "Sex"   # used for the gender-ratio column in the Summary sheet


INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "Fall 2026 Orientation Group Assignment FY 20260526-100456.xlsx"
OUTPUT_FILE = "orientation_groups_49.xlsx"


# ─────────────────────────────────────────
# NORMALIZATION HELPERS
# ─────────────────────────────────────────

# Full state name → 2-letter code (used as a fallback if geomarket prefix is missing)
STATE_NAME_TO_CODE = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR",
    "CALIFORNIA": "CA", "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID",
    "ILLINOIS": "IL", "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS",
    "KENTUCKY": "KY", "LOUISIANA": "LA", "MAINE": "ME", "MARYLAND": "MD",
    "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN", "MISSISSIPPI": "MS",
    "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK",
    "OREGON": "OR", "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC",
    "SOUTH DAKOTA": "SD", "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT",
    "VERMONT": "VT", "VIRGINIA": "VA", "WASHINGTON": "WA", "WEST VIRGINIA": "WV",
    "WISCONSIN": "WI", "WYOMING": "WY", "DISTRICT OF COLUMBIA": "DC",
}

USA_COUNTRY_VALUES = {
    "USA", "US", "U.S.", "U.S.A.",
    "UNITED STATES", "UNITED STATES OF AMERICA", "AMERICA",
}

# Country name normalization map (left side = uppercase variant seen in raw data)
COUNTRY_NORMALIZE = {
    "UNITED STATES OF AMERICA": "USA",
    "UNITED STATES": "USA",
    "U.S.A.": "USA",
    "U.S.": "USA",
    "US": "USA",
    "AMERICA": "USA",
    "UK": "UNITED KINGDOM",
    "U.K.": "UNITED KINGDOM",
    "GREAT BRITAIN": "UNITED KINGDOM",
    "PRC": "CHINA",
    "PEOPLE'S REPUBLIC OF CHINA": "CHINA",
    "ROC": "TAIWAN",
    "REPUBLIC OF KOREA": "SOUTH KOREA",
    "KOREA, SOUTH": "SOUTH KOREA",
    "KOREA SOUTH": "SOUTH KOREA",
    "VIET NAM": "VIETNAM",
    "BAHAMAS": "THE BAHAMAS",
    "TRINIDAD & TOBAGO": "TRINIDAD AND TOBAGO",
}


def normalize_state(raw):
    """
    Convert raw geomarket text (e.g. 'CT-03 Fairfield Co', 'INT-IN India',
    'Connecticut') to a 2-letter US state code, or '' if international/unknown.
    """
    if not raw:
        return ""
    s = str(raw).strip().upper()

    # International prefix means no US state
    if s.startswith("INT-") or s.startswith("INT "):
        return ""

    # Geomarket pattern: 'XX-NN Description'
    m = re.match(r"^([A-Z]{2})[- ]\d", s)
    if m:
        return m.group(1)

    # Bare 2-letter code
    if len(s) == 2 and s.isalpha():
        return s

    # Full state name
    if s in STATE_NAME_TO_CODE:
        return STATE_NAME_TO_CODE[s]

    # Sometimes geomarket starts with state code followed by a space then word
    m = re.match(r"^([A-Z]{2})\b", s)
    if m and m.group(1) in STATE_NAME_TO_CODE.values():
        return m.group(1)

    return ""


def normalize_country(raw):
    """Standardize country name to uppercase canonical form. 'USA' for United States."""
    if not raw:
        return ""
    c = str(raw).strip().upper()
    if c in USA_COUNTRY_VALUES:
        return "USA"
    return COUNTRY_NORMALIZE.get(c, c)


# ─────────────────────────────────────────
# STEP 1: LOAD & CLASSIFY
# ─────────────────────────────────────────
def load_and_classify(filepath):
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip()

    # Keep the original geomarket text for the output, but build a normalized state code
    df["_raw_state"] = df[COL_STATE].fillna("").astype(str).str.strip()
    df["_state_code"] = df["_raw_state"].apply(normalize_state)

    df[COL_COUNTRY] = df[COL_COUNTRY].fillna("").astype(str).apply(normalize_country)
    df[COL_HIGH_SCHOOL] = df[COL_HIGH_SCHOOL].fillna("").astype(str).str.strip().str.lower()
    df[COL_SPORT]       = df[COL_SPORT].fillna("").astype(str).str.strip().str.lower()
    df[COL_POSSE]       = df[COL_POSSE].fillna("").astype(str).str.strip().str.lower()

    # The Posse column in this dataset is a yes/no flag, NOT a cohort identifier.
    # Treat anything that doesn't indicate Posse membership as null so it isn't
    # used as a uniqueness key (otherwise every "no" student conflicts).
    NEGATIVE_FLAGS = {"", "no", "n", "false", "0", "none", "na", "n/a"}
    df["_is_posse"] = ~df[COL_POSSE].isin(NEGATIVE_FLAGS)
    # For uniqueness purposes, only use the value when it actually marks Posse.
    df["_posse_key"] = df.apply(
        lambda r: r[COL_POSSE] if r["_is_posse"] else "", axis=1
    )

    df["is_international"] = df[COL_COUNTRY].apply(lambda c: c != "" and c != "USA")
    df["is_northeast"]      = df["_state_code"].apply(lambda s: s in NORTHEAST_STATES)
    df["is_mact"]           = df["_state_code"].apply(lambda s: s in MACT_STATES)
    df["is_nonNE_domestic"] = (~df["is_international"]) & (~df["is_northeast"])

    return df


# ─────────────────────────────────────────
# STEP 2: SORT using PLACEMENT DIFFICULTY
# ─────────────────────────────────────────
def sort_by_difficulty(df):
    def priority(row):
        if row["_is_posse"]:            return 0
        if row["is_international"]:     return 1
        if row["is_nonNE_domestic"]:    return 2
        if row["is_mact"]:              return 3
        return 4

    df = df.copy()
    df["_priority"] = df.apply(priority, axis=1)
    df = df.sort_values("_priority", kind="stable")
    return df


# ─────────────────────────────────────────
# STEP 3: GROUP STATE
# ─────────────────────────────────────────
def empty_group():
    return {
        "students":    [],
        "total":       0,
        "ne":          0,
        "mact":        0,
        "nonNE_dom":   0,
        "intl":        0,
        "states":      set(),
        "countries":   set(),
        "highschools": set(),
        "sports":      set(),
        "posses":      set(),
    }




# ─────────────────────────────────────────
# STEP 4: CONSTRAINT CHECK
# ─────────────────────────────────────────
def can_assign(student, group):
    if group["total"] >= MAX_GROUP_SIZE:
        return False
    if student["is_northeast"] and group["ne"] >= MAX_NE_PER_GROUP:
        return False
    if student["is_mact"] and group["mact"] >= MAX_MACT_PER_GROUP:
        return False
    if student["is_nonNE_domestic"] and group["nonNE_dom"] >= MAX_NON_NE_DOM_PER_GROUP:
        return False
    if student["is_international"] and group["intl"] >= MAX_INTL_PER_GROUP:
        return False
    state_code = student["_state_code"]
    if student["is_nonNE_domestic"] and state_code and state_code in group["states"]:
        return False
    if student["is_international"] and student[COL_COUNTRY] and student[COL_COUNTRY] in group["countries"]:
        return False
    hs = student[COL_HIGH_SCHOOL]
    if hs and hs in group["highschools"]:
        return False
    sport = student[COL_SPORT]
    if sport and sport in group["sports"]:
        return False
    posse = student["_posse_key"]
    if posse and posse in group["posses"]:
        return False
    return True


def assign_student(student, group):
    group["students"].append(student.name)   # DataFrame index
    group["total"]  += 1
    if student["is_northeast"]:      group["ne"]        += 1
    if student["is_mact"]:           group["mact"]      += 1
    if student["is_nonNE_domestic"]: group["nonNE_dom"] += 1
    if student["is_international"]:  group["intl"]      += 1
    if student["is_nonNE_domestic"] and student["_state_code"]:
        group["states"].add(student["_state_code"])
    if student["is_international"] and student[COL_COUNTRY]:
        group["countries"].add(student[COL_COUNTRY])
    if student[COL_HIGH_SCHOOL]: group["highschools"].add(student[COL_HIGH_SCHOOL])
    if student[COL_SPORT]:       group["sports"].add(student[COL_SPORT])
    if student["_posse_key"]:    group["posses"].add(student["_posse_key"])


def remove_student(student, group):
    group["students"].remove(student.name)
    group["total"]  -= 1
    if student["is_northeast"]:      group["ne"]        -= 1
    if student["is_mact"]:           group["mact"]      -= 1
    if student["is_nonNE_domestic"]: group["nonNE_dom"] -= 1
    if student["is_international"]:  group["intl"]      -= 1
    # discard (no-op if absent)
    group["states"].discard(student["_state_code"])
    group["countries"].discard(student[COL_COUNTRY])
    group["highschools"].discard(student[COL_HIGH_SCHOOL])
    group["sports"].discard(student[COL_SPORT])
    group["posses"].discard(student["_posse_key"])


# ─────────────────────────────────────────
# STEP 5: GREEDY ASSIGNMENT
# ─────────────────────────────────────────
def greedy_assign(df, groups):
    random.seed(RANDOM_SEED)
    unassigned = []
    group_order = list(range(NUM_GROUPS))

    for idx, student in df.iterrows():
        # Bias toward the smallest groups, then shuffle ties so we don't
        # always fill group 1 first.
        random.shuffle(group_order)
        group_order_sorted = sorted(group_order, key=lambda g: groups[g]["total"])
        placed = False
        for g in group_order_sorted:
            if can_assign(student, groups[g]):
                assign_student(student, groups[g])
                placed = True
                break
        if not placed:
            unassigned.append(idx)

    return unassigned
def fix_international_minimums(df, groups):
    """
    Enforce the rule: each group has either 0 internationals, OR between
    MIN_INTL_PER_GROUP and MAX_INTL_PER_GROUP. A group with exactly 1
    international student is invalid.

    Strategy for each invalid group (0 < intl < MIN_INTL_PER_GROUP):
      1. Try to *raise* it to MIN by pulling an international from a donor
         group that has > MIN (so the donor stays valid).
      2. If that fails, *drain* it to 0 by pushing its lone international
         into a donor group that currently has ≥ MIN-1 (so the donor lands
         in the valid 2–3 range) — only if the receiving group can accept.
    """
    def try_raise(g_idx, group):
        for h_idx, donor in enumerate(groups):
            if h_idx == g_idx or donor["intl"] <= MIN_INTL_PER_GROUP:
                continue
            for s_idx in donor["students"][:]:
                student = df.loc[s_idx]
                if not student["is_international"]:
                    continue
                remove_student(student, donor)
                if can_assign(student, group):
                    assign_student(student, group)
                    return True
                assign_student(student, donor)  # restore
        return False

    def try_drain(g_idx, group):
        # Move every international out of this group into a group that
        # currently has between MIN-1 and MAX-1 internationals (so it stays
        # valid after gaining one).
        for s_idx in group["students"][:]:
            student = df.loc[s_idx]
            if not student["is_international"]:
                continue
            remove_student(student, group)
            placed = False
            for h_idx, receiver in enumerate(groups):
                if h_idx == g_idx:
                    continue
                if not (MIN_INTL_PER_GROUP - 1 <= receiver["intl"] < MAX_INTL_PER_GROUP):
                    continue
                if can_assign(student, receiver):
                    assign_student(student, receiver)
                    placed = True
                    break
            if not placed:
                # Put her back; we can't fully drain
                assign_student(student, group)
                return False
        return True

    for g_idx, group in enumerate(groups):
        attempts = 0
        while 0 < group["intl"] < MIN_INTL_PER_GROUP and attempts < NUM_GROUPS:
            attempts += 1
            if try_raise(g_idx, group):
                continue
            if try_drain(g_idx, group):
                break
            # Neither move possible — leave it for validation to flag
            break



# ─────────────────────────────────────────
# STEP 6a: NON-NE DOMESTIC OVERFLOW FIX
# If any group is over MAX_NON_NE_DOM_PER_GROUP, try to relocate one of its
# non-NE-domestic students into another group that still has room.
# Pure swap — does not bend constraints.
# ─────────────────────────────────────────
def try_swap_nonNE(df, groups):
    """
    For groups over the non-NE domestic cap, attempt a swap to resolve.
    Also handles state-duplicate soft-placements by trying a two-level swap.
    """
    changed = True
    while changed:
        changed = False
        for g_idx, group in enumerate(groups):
            if group["nonNE_dom"] <= MAX_NON_NE_DOM_PER_GROUP:
                continue
            # Try moving each non-NE domestic student out of the overfull group
            for s_idx in group["students"][:]:
                student = df.loc[s_idx]
                if not student["is_nonNE_domestic"]:
                    continue
                remove_student(student, group)
                placed = False
                for h_idx, receiver in enumerate(groups):
                    if h_idx == g_idx:
                        continue
                    if receiver["nonNE_dom"] >= MAX_NON_NE_DOM_PER_GROUP:
                        continue
                    if can_assign(student, receiver):
                        assign_student(student, receiver)
                        print(f"   ↳ Swapped {df.loc[s_idx]['First']} {df.loc[s_idx]['Last']} "
                              f"Group {g_idx+1} → Group {h_idx+1} (nonNE_dom fix)")
                        placed = True
                        changed = True
                        break
                if not placed:
                    assign_student(student, group)  # restore
                if group["nonNE_dom"] <= MAX_NON_NE_DOM_PER_GROUP:
                    break


def chain_swap_dup_state(df, groups, max_passes=5):
    """
    Resolve duplicate non-NE state violations using a 2-hop chain swap.

    For each group G_src that has duplicate state S:
      1. Try a direct 1-hop move (same as fix_dup_state_via_swap).
      2. If that fails, look for an intermediate group G_mid that has room
         for one of G_src's duplicate-state students AND a current member Y
         we can re-home into a third group G_dst (cleanly). Then:
            - Move Y from G_mid → G_dst
            - Move X from G_src → G_mid
         The net effect is: G_src loses a duplicate, G_mid stays balanced,
         G_dst gains one student. No constraint is bent.

    Repeats until no progress is made or max_passes is hit.
    """
    for _ in range(max_passes):
        changed = False
        for g_src, group in enumerate(groups):
            # Find a duplicate state in this group
            state_counts = {}
            for s_idx in group["students"]:
                s = df.loc[s_idx]
                if s["is_nonNE_domestic"] and s["_state_code"]:
                    state_counts[s["_state_code"]] = state_counts.get(s["_state_code"], 0) + 1
            dup_states = [st for st, c in state_counts.items() if c > 1]
            if not dup_states:
                continue

            for dup_state in dup_states:
                # Collect duplicate students in source group
                candidates_x = [
                    s_idx for s_idx in group["students"][:]
                    if df.loc[s_idx]["is_nonNE_domestic"]
                    and df.loc[s_idx]["_state_code"] == dup_state
                ]
                resolved = False
                for x_idx in candidates_x:
                    X = df.loc[x_idx]

                    # ── Try a 1-hop direct move first ──
                    remove_student(X, group)
                    for g_dst, dst in enumerate(groups):
                        if g_dst == g_src:
                            continue
                        if can_assign(X, dst):
                            assign_student(X, dst)
                            print(f"   ↳ Direct-moved {X['First']} {X['Last']} "
                                  f"Group {g_src+1} → Group {g_dst+1} "
                                  f"(dup {dup_state} fix)")
                            resolved = True
                            changed = True
                            break
                    if resolved:
                        break

                    # ── 2-hop chain swap ──
                    for g_mid, mid in enumerate(groups):
                        if g_mid == g_src:
                            continue
                        # G_mid must currently *not* accept X (otherwise 1-hop would have worked)
                        # but should become acceptable if we remove a current member Y.
                        for y_idx in mid["students"][:]:
                            Y = df.loc[y_idx]
                            # Don't shuffle posse / international students lightly
                            remove_student(Y, mid)
                            if not can_assign(X, mid):
                                assign_student(Y, mid)
                                continue
                            # X could go to G_mid if we re-home Y. Find a third group.
                            placed_y = False
                            for g_dst, dst in enumerate(groups):
                                if g_dst in (g_src, g_mid):
                                    continue
                                if can_assign(Y, dst):
                                    assign_student(Y, dst)
                                    assign_student(X, mid)
                                    print(f"   ↳ Chain-swapped: {X['First']} {X['Last']} "
                                          f"Group {g_src+1} → Group {g_mid+1}, "
                                          f"{Y['First']} {Y['Last']} "
                                          f"Group {g_mid+1} → Group {g_dst+1} "
                                          f"(dup {dup_state} fix)")
                                    placed_y = True
                                    break
                            if placed_y:
                                resolved = True
                                changed = True
                                break
                            # restore Y and try the next member
                            assign_student(Y, mid)
                        if resolved:
                            break
                    if resolved:
                        break

                    # Couldn't resolve via this X — restore it and try the next
                    assign_student(X, group)

                if resolved:
                    break  # re-scan from the top after a change
            if changed:
                break
        if not changed:
            return


def fix_dup_state_via_swap(df, groups, unassigned):

    """
    After soft-placement, if a student landed with a duplicate state,
    try a chain swap: find someone in that group with a different state
    who can move elsewhere, freeing a clean slot.
    """
    for idx in list(unassigned):
        # unassigned is already empty at this point; we check violations instead
        pass

    # Check all groups for duplicate non-NE states
    for g_idx, group in enumerate(groups):
        state_counts = {}
        for s_idx in group["students"]:
            s = df.loc[s_idx]
            if s["is_nonNE_domestic"] and s["_state_code"]:
                state_counts[s["_state_code"]] = state_counts.get(s["_state_code"], 0) + 1

        for dup_state, count in state_counts.items():
            if count < 2:
                continue
            # Find one of the duplicates to move out
            for s_idx in group["students"][:]:
                student = df.loc[s_idx]
                if not student["is_nonNE_domestic"] or student["_state_code"] != dup_state:
                    continue
                remove_student(student, group)
                placed = False
                for h_idx, receiver in enumerate(groups):
                    if h_idx == g_idx:
                        continue
                    if receiver["nonNE_dom"] >= MAX_NON_NE_DOM_PER_GROUP:
                        continue
                    if can_assign(student, receiver):
                        assign_student(student, receiver)
                        print(f"   ↳ Moved {df.loc[s_idx]['First']} {df.loc[s_idx]['Last']} "
                              f"Group {g_idx+1} → Group {h_idx+1} (dup state {dup_state} fix)")
                        placed = True
                        break
                if not placed:
                    assign_student(student, group)  # restore
                else:
                    break  # one move per duplicate per pass; re-validate after



# ─────────────────────────────────────────
# STEP 6c: SOFT-RELAXATION PASS

# Place any remaining unassigned students into the *least disruptive* group
# by scoring each group's constraint violations and picking the lowest.
# ─────────────────────────────────────────
# Penalty weights: lower = softer constraint to break.
_PENALTY = {
    "size_hard":   10_000,   # never exceed MAX_GROUP_SIZE — effectively forbidden
    "ne_cap":         60,
    "mact_cap":       50,
    "intl_cap":       40,
    "dup_hs":         35,
    "dup_state":      30,
    "dup_country":    30,
    "dup_posse":      25,
    "dup_sport":      20,
    "nonNE_cap":  10_000,    # hard — same as size_hard
}


def soft_score(student, group):
    """Return (total_penalty, [reasons]) for placing `student` in `group`."""
    reasons, total = [], 0.0
    def add(key, label):
        nonlocal total
        total += _PENALTY[key]
        reasons.append((_PENALTY[key], label))
    if group["total"] >= MAX_GROUP_SIZE:
        add("size_hard", "group at MAX_GROUP_SIZE")
    if student["is_northeast"]      and group["ne"]        >= MAX_NE_PER_GROUP:        add("ne_cap",   "NE cap")
    if student["is_mact"]           and group["mact"]      >= MAX_MACT_PER_GROUP:      add("mact_cap", "MA/CT cap")
    if student["is_international"]  and group["intl"]      >= MAX_INTL_PER_GROUP:      add("intl_cap", "intl cap")
    if student["is_nonNE_domestic"] and group["nonNE_dom"] >= MAX_NON_NE_DOM_PER_GROUP:
        add("nonNE_cap", f"non-NE-dom cap (currently {group['nonNE_dom']})")
    if student["is_nonNE_domestic"] and student["_state_code"] and student["_state_code"] in group["states"]:
        add("dup_state", f"duplicate state {student['_state_code']}")
    if student["is_international"] and student[COL_COUNTRY] and student[COL_COUNTRY] in group["countries"]:
        add("dup_country", f"duplicate country {student[COL_COUNTRY]}")
    if student[COL_HIGH_SCHOOL] and student[COL_HIGH_SCHOOL] in group["highschools"]:
        add("dup_hs", "duplicate high school")
    if student[COL_SPORT] and student[COL_SPORT] in group["sports"]:
        add("dup_sport", f"duplicate sport ({student[COL_SPORT]})")
    if student["_posse_key"] and student["_posse_key"] in group["posses"]:
        add("dup_posse", "duplicate Posse cohort")
    # tiebreaker: prefer smaller groups
    total += group["total"] * 0.01
    return total, reasons


def soft_place_remaining(df, unassigned, groups):
    """Place every still-unassigned student in their best-fit group, with a
    printed justification. Returns the list of (student_idx, group_idx, reasons)."""
    log = []
    for idx in list(unassigned):
        stu = df.loc[idx]
        best = min(((soft_score(stu, g), gi) for gi, g in enumerate(groups)),
                   key=lambda x: x[0][0])
        (score, reasons), gi = best
        if score >= _PENALTY["size_hard"]:
            # Group is literally full — keep her unassigned
            continue
        assign_student(stu, groups[gi])
        unassigned.remove(idx)
        log.append((idx, gi, reasons))
        name = f"{stu['First']} {stu['Last']}"
        print(f"   ↳ Soft-placed {name} into Group {gi+1} (score {score:.2f})")
        if reasons:
            for w, label in reasons:
                print(f"        - {label}  (+{w})")
        else:
            print(f"        - fits cleanly (no violations)")
    return log


# ─────────────────────────────────────────
# STEP 6b: VALIDATION
# ─────────────────────────────────────────
def validate_groups(df, groups):
    """Re-check every group against the constraints; print/return violations."""
    violations = []
    for g_num, group in enumerate(groups, start=1):
        if group["total"] > MAX_GROUP_SIZE:
            violations.append(f"Group {g_num}: total={group['total']} > {MAX_GROUP_SIZE}")
        if group["ne"] > MAX_NE_PER_GROUP:
            violations.append(f"Group {g_num}: NE={group['ne']} > {MAX_NE_PER_GROUP}")
        if group["mact"] > MAX_MACT_PER_GROUP:
            violations.append(f"Group {g_num}: MA/CT={group['mact']} > {MAX_MACT_PER_GROUP}")
        if group["nonNE_dom"] > MAX_NON_NE_DOM_PER_GROUP:
            violations.append(f"Group {g_num}: nonNE_dom={group['nonNE_dom']} > {MAX_NON_NE_DOM_PER_GROUP}")
        if group["intl"] > MAX_INTL_PER_GROUP:
            violations.append(f"Group {g_num}: intl={group['intl']} > {MAX_INTL_PER_GROUP}")
        # Rule: intl must be 0 OR in [MIN, MAX]. Exactly 1 (or any 1..MIN-1) is invalid.
        if 0 < group["intl"] < MIN_INTL_PER_GROUP:
            violations.append(
                f"Group {g_num}: intl={group['intl']} (must be 0 or "
                f"{MIN_INTL_PER_GROUP}–{MAX_INTL_PER_GROUP})"
            )


        # uniqueness checks
        seen = {"nonNE_state": [], "intl_country": [], "hs": [], "sport": [], "posse": []}
        for s_idx in group["students"]:
            s = df.loc[s_idx]
            if s["is_nonNE_domestic"] and s["_state_code"]:
                seen["nonNE_state"].append(s["_state_code"])
            if s["is_international"] and s[COL_COUNTRY]:
                seen["intl_country"].append(s[COL_COUNTRY])
            if s[COL_HIGH_SCHOOL]:
                seen["hs"].append(s[COL_HIGH_SCHOOL])
            if s[COL_SPORT]:
                seen["sport"].append(s[COL_SPORT])
            if s["_posse_key"]:
                seen["posse"].append(s["_posse_key"])
        for key, vals in seen.items():
            dupes = [v for v in set(vals) if vals.count(v) > 1]
            if dupes:
                violations.append(f"Group {g_num}: duplicate {key} -> {dupes}")
    return violations


# ─────────────────────────────────────────
# STEP 7: OUTPUT TO EXCEL
# ─────────────────────────────────────────
INTERNAL_COLS = ["_priority", "_raw_state", "_state_code", "_is_posse", "_posse_key",
                 "is_international", "is_northeast", "is_mact", "is_nonNE_domestic"]


def write_output(df, groups, unassigned, violations):
    # Restore the original geomarket text in the output
    df = df.copy()
    if "_raw_state" in df.columns:
        df[COL_STATE] = df["_raw_state"]

    df["Group Number"] = None
    for g_num, group in enumerate(groups, start=1):
        for s_idx in group["students"]:
            df.at[s_idx, "Group Number"] = g_num

    drop_cols = INTERNAL_COLS

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Master sheet — full roster with Group Number
        master_df = df.drop(columns=drop_cols, errors="ignore").sort_values(
            by=["Group Number"], na_position="last"
        )
        master_df.to_excel(writer, sheet_name="Master", index=False)

        # ── Gender helpers ─────────────────────────────────────
        # The Sex column should be 'F' / 'M' for everyone. Anything else (blank,
        # weird casing, free-text Gender Identity etc.) is bucketed as "Other".
        def _sex_label(raw):
            if raw is None:
                return "Other"
            s = str(raw).strip().upper()
            if s in ("F", "FEMALE"):
                return "F"
            if s in ("M", "MALE"):
                return "M"
            return "Other"

        def gender_counts(student_indices):
            f = m = o = 0
            for s_idx in student_indices:
                val = df.loc[s_idx].get(COL_SEX, "") if COL_SEX in df.columns else ""
                lab = _sex_label(val)
                if   lab == "F": f += 1
                elif lab == "M": m += 1
                else:            o += 1
            return f, m, o

        def fmt_ratio(f, m):
            # "F:M ratio (F%)" — e.g. "7:6 (54% F)"
            tot = f + m
            if tot == 0:
                return "—"
            pct = round(100 * f / tot)
            return f"{f}:{m} ({pct}% F)"

        # Summary sheet
        summary_rows = []
        all_f = all_m = all_o = 0
        for g_num, group in enumerate(groups, start=1):
            f, m, o = gender_counts(group["students"])
            all_f += f; all_m += m; all_o += o
            summary_rows.append({
                "Group":            g_num,
                "Total Students":   group["total"],
                "NE Students":      group["ne"],
                "MA/CT Students":   group["mact"],
                "Non-NE Domestic":  group["nonNE_dom"],
                "International":    group["intl"],
                "Female":           f,
                "Male":             m,
                "Other / Unknown":  o,
                "F:M Ratio":        fmt_ratio(f, m),
                "Intl OK":          "✓" if (group["intl"] == 0 or group["intl"] >= MIN_INTL_PER_GROUP) else "⚠ INVALID (1)",
                "Size OK":          "✓" if MIN_GROUP_SIZE <= group["total"] <= MAX_GROUP_SIZE else "⚠",
            })
        # Totals row across the whole cohort
        summary_rows.append({
            "Group":            "TOTAL",
            "Total Students":   sum(g["total"]    for g in groups),
            "NE Students":      sum(g["ne"]       for g in groups),
            "MA/CT Students":   sum(g["mact"]     for g in groups),
            "Non-NE Domestic":  sum(g["nonNE_dom"] for g in groups),
            "International":    sum(g["intl"]    for g in groups),
            "Female":           all_f,
            "Male":             all_m,
            "Other / Unknown":  all_o,
            "F:M Ratio":        fmt_ratio(all_f, all_m),
            "Intl OK":          "",
            "Size OK":          "",
        })
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)


        # One sheet per group
        for g_num, group in enumerate(groups, start=1):
            group_df = df.loc[group["students"]].drop(
                columns=drop_cols + ["Group Number"], errors="ignore"
            )
            group_df.to_excel(writer, sheet_name=f"Group {g_num}", index=False)

        # Unassigned sheet
        if unassigned:
            unassigned_df = df.loc[unassigned].drop(columns=drop_cols, errors="ignore")
            unassigned_df.to_excel(writer, sheet_name="Unassigned", index=False)

        # Violations sheet
        if violations:
            pd.DataFrame({"Violation": violations}).to_excel(
                writer, sheet_name="Violations", index=False
            )

    # Formatting pass
    wb = load_workbook(OUTPUT_FILE)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    wb.save(OUTPUT_FILE)

    print(f"\n Done! Output saved to {OUTPUT_FILE}")
    print(f"   Groups created:       {NUM_GROUPS}")
    print(f"   Students assigned:    {sum(g['total'] for g in groups)}")
    print(f"   Unassigned (manual):  {len(unassigned)}")
    intl_warnings = [i+1 for i, g in enumerate(groups) if 0 < g["intl"] < MIN_INTL_PER_GROUP]
    if intl_warnings:
        print(f"   ⚠ Groups with an invalid intl count (must be 0 or "
              f"{MIN_INTL_PER_GROUP}–{MAX_INTL_PER_GROUP}): {intl_warnings}")

    if violations:
        print(f"   ⚠ {len(violations)} validation issue(s) — see 'Violations' sheet")


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────
if __name__ == "__main__":
    print(f"Loading {INPUT_FILE}...")
    df = load_and_classify(INPUT_FILE)
    print(f"  {len(df)} students loaded")
    print(f"  International: {int(df['is_international'].sum())}")
    print(f"  Northeast:     {int(df['is_northeast'].sum())}")
    print(f"  MA/CT:         {int(df['is_mact'].sum())}")
    print(f"  Non-NE dom:    {int(df['is_nonNE_domestic'].sum())}")

    # ── Capacity sanity check ──────────────────────────────────
    total_students = len(df)
    capacity = NUM_GROUPS * MAX_GROUP_SIZE
    if total_students > capacity:
        needed_groups = -(-total_students // MAX_GROUP_SIZE)         # ceiling div
        needed_size   = -(-total_students // NUM_GROUPS)
        print()
        print(f"⚠ CAPACITY WARNING: {total_students} students but only "
              f"{NUM_GROUPS} groups × {MAX_GROUP_SIZE} = {capacity} slots.")
        print(f"   To fit everyone you need either:")
        print(f"     - NUM_GROUPS      ≥ {needed_groups}  (at current MAX_GROUP_SIZE={MAX_GROUP_SIZE}), OR")
        print(f"     - MAX_GROUP_SIZE  ≥ {needed_size}  (at current NUM_GROUPS={NUM_GROUPS})")
        print(f"   Proceeding anyway; overflow students will land in the 'Unassigned' sheet.")
        print()

    df = sort_by_difficulty(df)
    groups = [empty_group() for _ in range(NUM_GROUPS)]

    print("Running greedy assignment...")
    unassigned = greedy_assign(df, groups)

    print("Fixing international minimums...")
    fix_international_minimums(df, groups)

    # ── NEW: fix nonNE_dom overflows before soft-placing ──
    print("Fixing non-NE domestic overflows...")
    try_swap_nonNE(df, groups)

    if unassigned:
        print(f"Soft-placing {len(unassigned)} remaining student(s) into best-fit groups...")
        soft_place_remaining(df, unassigned, groups)

    # ── NEW: clean up any duplicate-state side effects from soft-placement ──
    print("Fixing duplicate-state placements via 1-hop swap...")
    fix_dup_state_via_swap(df, groups, unassigned)

    # ── NEW: 2-hop chain swap for any duplicate state the 1-hop pass couldn't fix
    print("Fixing remaining duplicate-state placements via chain swap...")
    chain_swap_dup_state(df, groups)

    print("Validating...")


    violations = validate_groups(df, groups)
    for v in violations:
        print(f"   ⚠ {v}")

    print("\nPer-group summary:")
    for g_num, g in enumerate(groups, start=1):
        print(f"  Group {g_num:2d}: total={g['total']:2d}  NE={g['ne']:2d}  "
              f"MA/CT={g['mact']:2d}  nonNE_dom={g['nonNE_dom']:2d}  intl={g['intl']:2d}")

    print("\nWriting output...")
    write_output(df, groups, unassigned, violations)
